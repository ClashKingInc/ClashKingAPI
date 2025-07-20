"""
War statistics Excel export functionality
"""
from datetime import datetime
from tempfile import NamedTemporaryFile
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook

from routers.v2.exports.utils import (
    add_clashking_logo_to_sheet,
    format_table,
    add_title_to_sheet,
    add_filter_summary_to_sheet,
    add_table_headers,
    add_summary_data,
    add_section_title
)


class WarStatsExporter:
    """Excel exporter for war statistics"""
    
    def __init__(self, war_hits: List[Dict[str, Any]], player_tag: str):
        self.war_hits = war_hits
        self.player_tag = player_tag
        self.wb = Workbook()
        
        # Calculate basic statistics
        self.total_attacks = len(war_hits)
        self.total_stars = sum(hit.get("stars", 0) for hit in war_hits)
        self.total_destruction = sum(hit.get("destruction_percentage", 0) for hit in war_hits)
        self.total_wars = len(set(hit.get("war_tag", "") for hit in war_hits if hit.get("war_tag")))
        
        # Star distribution
        self.three_stars = sum(1 for hit in war_hits if hit.get("stars") == 3)
        self.two_stars = sum(1 for hit in war_hits if hit.get("stars") == 2)
        self.one_star = sum(1 for hit in war_hits if hit.get("stars") == 1)
        self.zero_stars = sum(1 for hit in war_hits if hit.get("stars") == 0)
        
        # TH analysis
        self.th_attacks = {}
        self.opponent_th_attacks = {}
        self.th_stars_breakdown = {}
        self.attack_matchups = {}
        self.defense_matchups = {}
        
        # Calculate current player TH and matchup data
        self._analyze_th_data()
    
    def _analyze_th_data(self):
        """Analyze TH level data and matchups"""
        for hit in self.war_hits:
            own_th = hit.get("attacker_townhall", "Unknown")
            enemy_th = hit.get("defender_townhall", "Unknown")
            stars = hit.get("stars", 0)
            destruction = hit.get("destruction_percentage", 0)

            self.th_attacks[own_th] = self.th_attacks.get(own_th, 0) + 1
            self.opponent_th_attacks[enemy_th] = self.opponent_th_attacks.get(enemy_th, 0) + 1

            # Track stars by TH level
            if own_th not in self.th_stars_breakdown:
                self.th_stars_breakdown[own_th] = {0: 0, 1: 0, 2: 0, 3: 0}
            self.th_stars_breakdown[own_th][stars] = self.th_stars_breakdown[own_th].get(stars, 0) + 1

            # Track attack matchups (when attacking)
            if own_th != "Unknown" and enemy_th != "Unknown":
                matchup_key = (own_th, enemy_th)
                if matchup_key not in self.attack_matchups:
                    self.attack_matchups[matchup_key] = {0: 0, 1: 0, 2: 0, 3: 0, 'total_destruction': 0, 'attack_count': 0}
                self.attack_matchups[matchup_key][stars] = self.attack_matchups[matchup_key].get(stars, 0) + 1
                self.attack_matchups[matchup_key]['total_destruction'] += destruction
                self.attack_matchups[matchup_key]['attack_count'] += 1

                # Track defense matchups (when being attacked) - reverse perspective
                defense_key = (enemy_th, own_th)  # defender TH, attacker TH
                if defense_key not in self.defense_matchups:
                    self.defense_matchups[defense_key] = {0: 0, 1: 0, 2: 0, 3: 0, 'total_destruction': 0, 'attack_count': 0}
                self.defense_matchups[defense_key][stars] = self.defense_matchups[defense_key].get(stars, 0) + 1
                self.defense_matchups[defense_key]['total_destruction'] += destruction
                self.defense_matchups[defense_key]['attack_count'] += 1

        # Determine current player TH (highest TH level used in attacks)
        self.current_player_th = max((int(th) for th in self.th_attacks.keys() if str(th).isdigit()), default=0)
    
    async def _create_summary_sheet(self, filter_summary: str) -> None:
        """Create the summary sheet with statistics and analysis"""
        ws_summary = self.wb.active
        ws_summary.title = "Summary"
        
        # Add logo
        await add_clashking_logo_to_sheet(ws_summary, anchor_cell="D1")
        
        # Title and filters
        add_title_to_sheet(ws_summary, f"War Statistics Summary for {self.player_tag}", "A1:B1")
        add_filter_summary_to_sheet(ws_summary, filter_summary, "A1:B1")
        
        # Basic statistics
        summary_data = self._get_summary_data()
        ws_summary.append(["Statistic", "Value"])
        start_summary_row = ws_summary.max_row
        add_summary_data(ws_summary, summary_data, preserve_numeric_types=True)
        end_summary_row = ws_summary.max_row
        format_table(ws_summary, start_summary_row, end_summary_row, "Summary")

        # TH Attack Performance Table
        if self.th_stars_breakdown:
            self._add_th_performance_table(ws_summary)
    
    def _get_summary_data(self) -> List[Tuple[str, Any]]:
        """Get summary statistics data"""
        # War type breakdown
        cwl_attacks = sum(1 for hit in self.war_hits if hit.get("war_type") == "CWL")
        regular_attacks = sum(1 for hit in self.war_hits if hit.get("war_type") == "Regular")
        friendly_war_attacks = sum(1 for hit in self.war_hits if hit.get("war_type") not in ["CWL", "Regular"])
        
        # Fresh vs cleanup attacks
        fresh_attacks = sum(1 for hit in self.war_hits if hit.get("fresh_attack", False))
        cleanup_attacks = self.total_attacks - fresh_attacks
        
        # Destruction analysis
        high_destruction = sum(1 for hit in self.war_hits if hit.get("destruction_percentage", 0) >= 90)
        medium_destruction = sum(1 for hit in self.war_hits if 50 <= hit.get("destruction_percentage", 0) < 90)
        low_destruction = sum(1 for hit in self.war_hits if hit.get("destruction_percentage", 0) < 50)
        
        # Most common and targeted TH
        most_common_th = max(self.th_attacks.items(), key=lambda x: x[1]) if self.th_attacks else ("Unknown", 0)
        most_targeted_th = max(self.opponent_th_attacks.items(), key=lambda x: x[1]) if self.opponent_th_attacks else ("Unknown", 0)
        
        # Performance rates
        perfect_rate = round(self.three_stars / self.total_attacks * 100, 2) if self.total_attacks > 0 else 0
        fail_rate = round(self.zero_stars / self.total_attacks * 100, 2) if self.total_attacks > 0 else 0
        
        return [
            # Top Level Stats
            ["Total Wars", self.total_wars],
            ["Total Attacks", self.total_attacks],
            ["Total Stars", self.total_stars],
            ["", ""],  # Separator
            
            # War Stats
            ["CWL Attacks", f"{cwl_attacks} ({round(cwl_attacks / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["Regular War Attacks", f"{regular_attacks} ({round(regular_attacks / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["Friendly War Attacks", f"{friendly_war_attacks} ({round(friendly_war_attacks / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["", ""],  # Separator
            
            # Average Stats
            ["Average Stars", round(self.total_stars / self.total_attacks, 2) if self.total_attacks > 0 else 0],
            ["Average Destruction", f"{round(self.total_destruction / self.total_attacks, 2)}%" if self.total_attacks > 0 else "0%"],
            ["", ""],  # Separator
            
            # Attack Distribution
            ["3 Star Attacks", f"{self.three_stars} ({perfect_rate}%)"],
            ["2 Star Attacks", f"{self.two_stars} ({round(self.two_stars / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["1 Star Attacks", f"{self.one_star} ({round(self.one_star / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["0 Star Attacks", f"{self.zero_stars} ({fail_rate}%)"],
            ["", ""],  # Separator
            
            # TH Analysis
            ["Most Used TH", f"TH{most_common_th[0]} ({most_common_th[1]} attacks)"],
            ["Most Targeted TH", f"TH{most_targeted_th[0]} ({most_targeted_th[1]} attacks)"],
            ["", ""],  # Separator
            
            # Destruction Analysis
            ["High Destruction (≥90%)", f"{high_destruction} ({round(high_destruction / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["Medium Destruction (50-89%)", f"{medium_destruction} ({round(medium_destruction / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["Low Destruction (<50%)", f"{low_destruction} ({round(low_destruction / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["", ""],  # Separator
            
            # Attack Type
            ["Fresh Attacks", f"{fresh_attacks} ({round(fresh_attacks / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"],
            ["Cleanup Attacks", f"{cleanup_attacks} ({round(cleanup_attacks / self.total_attacks * 100, 2)}%)" if self.total_attacks > 0 else "0 (0%)"]
        ]
    
    def _add_th_performance_table(self, ws_summary):
        """Add TH Attack Performance Table"""
        add_section_title(ws_summary, "Attack Performance by Town Hall Level", "A:F")
        
        # Table headers
        headers = ["TH Level", "Total Attacks", "3⭐", "2⭐", "1⭐", "0⭐"]
        th_table_start = add_table_headers(ws_summary, headers)
        
        # Track rows that should be highlighted (current player TH)
        current_th_rows = []
        
        # Sort TH levels numerically
        for th in sorted(self.th_stars_breakdown.keys(), key=lambda x: int(x) if str(x).isdigit() else 999):
            if th == "Unknown":
                continue
            stars_data = self.th_stars_breakdown[th]
            total_th_attacks = sum(stars_data.values())
            if total_th_attacks > 0:
                ws_summary.append([
                    f"TH{th}",
                    total_th_attacks,
                    f"{stars_data[3]} ({round(stars_data[3] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[2]} ({round(stars_data[2] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[1]} ({round(stars_data[1] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[0]} ({round(stars_data[0] / total_th_attacks * 100, 1)}%)"
                ])
                
                # Mark current player TH row for highlighting
                if int(th) == self.current_player_th:
                    current_th_rows.append(ws_summary.max_row)
        
        th_table_end = ws_summary.max_row
        format_table(ws_summary, th_table_start, th_table_end, "THAttacks", highlight_rows=current_th_rows)
        
        # Attack Performance Matrix
        if self.attack_matchups:
            self._add_attack_performance_matrix(ws_summary)
        
        # Defense Performance Matrix
        if self.defense_matchups:
            self._add_defense_performance_matrix(ws_summary)
    
    def _add_attack_performance_matrix(self, ws_summary):
        """Add Attack Performance Matrix"""
        add_section_title(ws_summary, "Attack Performance Matrix")
        
        # Enhanced headers with efficiency indicators and destruction average
        headers = ["Matchup", "Attacks", "3⭐", "2⭐", "1⭐", "0⭐", "Avg Stars", "Avg Destruction", "Efficiency"]
        attack_table_start = add_table_headers(ws_summary, headers)
        
        # Track rows that should be highlighted (current player TH attacks)
        current_th_attack_rows = []
        
        # Sort by your TH first, then enemy TH
        for (your_th, enemy_th) in sorted(self.attack_matchups.keys(), key=lambda x: (
                int(x[0]) if str(x[0]).isdigit() else 999, int(x[1]) if str(x[1]).isdigit() else 999)):
            stars_data = self.attack_matchups[(your_th, enemy_th)]
            # Calculate total attacks excluding metadata keys
            total_matchup_attacks = stars_data.get('attack_count', sum(v for k, v in stars_data.items() if k in [0, 1, 2, 3]))
            if total_matchup_attacks > 0:
                # Calculate average stars
                avg_stars = (stars_data[3] * 3 + stars_data[2] * 2 + stars_data[1] * 1) / total_matchup_attacks

                # Calculate average destruction
                avg_destruction = stars_data.get('total_destruction', 0) / total_matchup_attacks if total_matchup_attacks > 0 else 0

                # Calculate efficiency based on 3 stars being perfect (100%)
                efficiency = f"{(avg_stars / 3.0 * 100):.0f}%"

                ws_summary.append([
                    f"TH{your_th} vs TH{enemy_th}",
                    total_matchup_attacks,
                    f"{stars_data[3]} ({round(stars_data[3] / total_matchup_attacks * 100, 1)}%)",
                    f"{stars_data[2]} ({round(stars_data[2] / total_matchup_attacks * 100, 1)}%)",
                    f"{stars_data[1]} ({round(stars_data[1] / total_matchup_attacks * 100, 1)}%)",
                    f"{stars_data[0]} ({round(stars_data[0] / total_matchup_attacks * 100, 1)}%)",
                    f"{avg_stars:.2f}",
                    f"{avg_destruction:.1f}%",
                    efficiency
                ])
                
                # Mark current player TH rows for highlighting
                if int(your_th) == self.current_player_th:
                    current_th_attack_rows.append(ws_summary.max_row)

        attack_table_end = ws_summary.max_row
        format_table(ws_summary, attack_table_start, attack_table_end, "AttackMatrix", highlight_rows=current_th_attack_rows)
    
    def _add_defense_performance_matrix(self, ws_summary):
        """Add Defense Performance Matrix"""
        add_section_title(ws_summary, "Defense Performance Matrix")
        
        # Detailed defense headers matching attack format
        headers = ["Matchup", "Defenses", "3⭐", "2⭐", "1⭐", "0⭐", "Avg Stars", "Avg Destruction", "Defense Rate"]
        defense_table_start = add_table_headers(ws_summary, headers)
        
        # Track rows that should be highlighted (current player TH defenses)
        current_th_defense_rows = []
        
        # Sort by enemy TH first, then your TH
        for (enemy_th, your_th) in sorted(self.defense_matchups.keys(), key=lambda x: (
                int(x[0]) if str(x[0]).isdigit() else 999, int(x[1]) if str(x[1]).isdigit() else 999)):
            stars_data = self.defense_matchups[(enemy_th, your_th)]
            # Calculate total attacks excluding metadata keys
            total_defense_attacks = stars_data.get('attack_count', sum(v for k, v in stars_data.items() if k in [0, 1, 2, 3]))
            if total_defense_attacks > 0:
                # Calculate average stars given up
                avg_stars_given = (stars_data[3] * 3 + stars_data[2] * 2 + stars_data[1] * 1) / total_defense_attacks

                # Calculate average destruction given up
                avg_destruction_given = stars_data.get('total_destruction', 0) / total_defense_attacks if total_defense_attacks > 0 else 0

                # Calculate defense rate (0 stars + partial holds)
                held_completely = stars_data[0]
                partial_holds = stars_data[1] + stars_data[2]
                defense_rate = f"{((held_completely + partial_holds) / total_defense_attacks * 100):.1f}%"

                ws_summary.append([
                    f"TH{enemy_th} vs TH{your_th}",
                    total_defense_attacks,
                    f"{stars_data[3]} ({round(stars_data[3] / total_defense_attacks * 100, 1)}%)",
                    f"{stars_data[2]} ({round(stars_data[2] / total_defense_attacks * 100, 1)}%)",
                    f"{stars_data[1]} ({round(stars_data[1] / total_defense_attacks * 100, 1)}%)",
                    f"{stars_data[0]} ({round(stars_data[0] / total_defense_attacks * 100, 1)}%)",
                    f"{avg_stars_given:.2f}",
                    f"{avg_destruction_given:.1f}%",
                    defense_rate
                ])
                
                # Mark current player TH rows for highlighting (defending TH)
                if int(your_th) == self.current_player_th:
                    current_th_defense_rows.append(ws_summary.max_row)

        defense_table_end = ws_summary.max_row
        format_table(ws_summary, defense_table_start, defense_table_end, "DefenseMatrix", highlight_rows=current_th_defense_rows)
    
    async def _create_attack_details_sheet(self, filter_summary: str) -> None:
        """Create the Attack Details sheet"""
        ws_attacks = self.wb.create_sheet(title="Attack Details")
        
        # Add logo
        await add_clashking_logo_to_sheet(ws_attacks)
        
        # Title and filters
        add_title_to_sheet(ws_attacks, f"Attack Details for {self.player_tag}")
        add_filter_summary_to_sheet(ws_attacks, filter_summary)
        
        # Headers
        headers = [
            "Player Name", "Player Tag", "TH Level", "War Date", "War Type",
            "Map Position", "Enemy Name", "Enemy Tag", "Enemy TH", "Enemy Map Position",
            "Stars", "Destruction %", "Attack Order", "Fresh Attack", "War Tag", "War State"
        ]
        start_row = add_table_headers(ws_attacks, headers)
        
        # Add data rows
        for hit in self.war_hits:
            ws_attacks.append([
                str(hit.get("attacker_name", "")),
                str(hit.get("attacker_tag", "")),
                str(hit.get("attacker_townhall", "")),
                str(hit.get("war_date", "")),
                str(hit.get("war_type", "")),
                str(hit.get("attacker_map_position", "")),
                str(hit.get("defender_name", "")),
                str(hit.get("defender_tag", "")),
                str(hit.get("defender_townhall", "")),
                str(hit.get("defender_map_position", "")),
                hit.get("stars", 0) or 0,
                hit.get("destruction_percentage", 0) or 0,
                str(hit.get("order", "")),
                "Yes" if hit.get("fresh_attack", False) else "No",
                str(hit.get("war_tag", "")),
                str(hit.get("war_state", ""))
            ])
        
        end_row = ws_attacks.max_row
        format_table(ws_attacks, start_row, end_row, "AttackStats")
    
    async def _create_defense_details_sheet(self, filter_summary: str) -> None:
        """Create the Defense Details sheet"""
        ws_defenses = self.wb.create_sheet(title="Defense Details")
        
        # Add logo
        await add_clashking_logo_to_sheet(ws_defenses)
        
        # Title and filters
        add_title_to_sheet(ws_defenses, f"Defense Details for {self.player_tag}")
        add_filter_summary_to_sheet(ws_defenses, filter_summary)
        
        # Headers for Defense Details
        headers = [
            "Defender Name", "Defender Tag", "Defender TH", "War Date", "War Type",
            "Map Position", "Attacker Name", "Attacker Tag", "Attacker TH", "Attacker Map Position",
            "Stars Given", "Destruction Given %", "Attack Order", "Fresh Attack", "War Tag", "War State"
        ]
        start_row = add_table_headers(ws_defenses, headers)
        
        # Add defense data rows (same data but from defender perspective)
        for hit in self.war_hits:
            ws_defenses.append([
                str(hit.get("defender_name", "")),
                str(hit.get("defender_tag", "")),
                str(hit.get("defender_townhall", "")),
                str(hit.get("war_date", "")),
                str(hit.get("war_type", "")),
                str(hit.get("defender_map_position", "")),
                str(hit.get("attacker_name", "")),
                str(hit.get("attacker_tag", "")),
                str(hit.get("attacker_townhall", "")),
                str(hit.get("attacker_map_position", "")),
                hit.get("stars", 0) or 0,
                hit.get("destruction_percentage", 0) or 0,
                str(hit.get("order", "")),
                "Yes" if hit.get("fresh_attack", False) else "No",
                str(hit.get("war_tag", "")),
                str(hit.get("war_state", ""))
            ])
        
        end_row = ws_defenses.max_row
        format_table(ws_defenses, start_row, end_row, "DefenseStats")
    
    async def create_excel_export(self, filter_summary: str) -> NamedTemporaryFile:
        """
        Create the complete Excel export with all sheets
        
        Args:
            filter_summary: Summary of applied filters
            
        Returns:
            NamedTemporaryFile containing the Excel workbook
        """
        # Create all sheets
        await self._create_summary_sheet(filter_summary)
        await self._create_attack_details_sheet(filter_summary)
        await self._create_defense_details_sheet(filter_summary)
        
        # Save to temporary file
        tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
        self.wb.save(tmp.name)
        tmp.seek(0)
        
        return tmp


async def export_player_war_stats_to_excel(
    war_hits: List[Dict[str, Any]], 
    player_tag: str, 
    filter_summary: str
) -> NamedTemporaryFile:
    """
    Export player war statistics to Excel format
    
    Args:
        war_hits: List of war hit data
        player_tag: Player tag for the export
        filter_summary: Summary of applied filters
        
    Returns:
        NamedTemporaryFile containing the Excel workbook
    """
    exporter = WarStatsExporter(war_hits, player_tag)
    return await exporter.create_excel_export(filter_summary)