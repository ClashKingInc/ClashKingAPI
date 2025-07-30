import tempfile
import uuid
import coc
from datetime import datetime

import aiohttp
from fastapi import HTTPException, APIRouter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as OpenpyxlImage
import pendulum as pend

from routers.v2.war.models import PlayerWarhitsFilter
from routers.v2.war.utils import collect_player_hits_from_wars
from utils.utils import fix_tag
from utils.database import MongoClient as mongo
from fastapi.responses import FileResponse

# Import the new export functionality
from routers.v2.exports.models import export_player_war_stats_to_excel

router = APIRouter(prefix="/v2/exports", tags=["Exports"], include_in_schema=True)


@router.get("/war/cwl-summary", name="Export CWL summary and members stats to Excel")
async def export_cwl_summary_to_excel(tag: str):
    def format_table(sheet, start_row, end_row, table_name_hint):
        table_name = f"{table_name_hint}_{uuid.uuid4().hex[:8]}"[:31]
        table_range = f"A{start_row}:Z{end_row}"

        try:
            table = Table(displayName=table_name, ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            table.tableStyleInfo = style
            sheet.add_table(table)
        except Exception as e:
            print(f"Failed to create table: {e}")

        # Manual styling as fallback
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if cell.row == start_row:
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    cell.fill = PatternFill(start_color="D7E3F0", end_color="D7E3F0", fill_type="solid")

    async def insert_logo_from_cdn(sheet, image_url: str, anchor_cell="A1", height=80):
        async with aiohttp.ClientSession() as session:
            async with session.get(image_url) as resp:
                if resp.status != 200:
                    raise Exception(f"Failed to download logo from CDN: {resp.status}")
                image_data = await resp.read()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
            tmp_img.write(image_data)
            tmp_img_path = tmp_img.name

        logo = OpenpyxlImage(tmp_img_path)
        logo.height = height
        logo.width = int(height * 3)
        sheet.add_image(logo, anchor_cell)

    def get_war_league_label(league_id):
        league_mapping = {
            48000000: "Unranked",
            48000001: "Bronze League III",
            48000002: "Bronze League II", 
            48000003: "Bronze League I",
            48000004: "Silver League III",
            48000005: "Silver League II",
            48000006: "Silver League I",
            48000007: "Gold League III",
            48000008: "Gold League II",
            48000009: "Gold League I",
            48000010: "Crystal League III",
            48000011: "Crystal League II",
            48000012: "Crystal League I",
            48000013: "Master League III",
            48000014: "Master League II",
            48000015: "Master League I",
            48000016: "Champion League III",
            48000017: "Champion League II", 
            48000018: "Champion League I"
        }
        return league_mapping.get(league_id, f"Unknown ({league_id})")

    try:
        clan_tag = fix_tag(tag)

        # Get current CWL season data
        pipeline = [
            {"$match": {"clan_tag": clan_tag}},
            {"$sort": {"season": -1}},
            {"$limit": 1}
        ]

        result = await mongo.clan_leaderboard.aggregate(pipeline).to_list(length=1)
        if not result:
            raise HTTPException(status_code=404, detail="No CWL data found for this clan")

        cwl_data = result[0]
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "CWL Summary"

        # Add logo
        await insert_logo_from_cdn(ws, "https://assets.clashk.ing/logos/crown-text-white-bg/BqlEp974170917vB1qK0zunfANJCGi0W031dTksEq7KQ9LoXWMFk0u77unHJa.png", "J1", 50)

        # Title
        ws["A1"] = f"CWL Summary for {cwl_data.get('clan_name', 'Unknown')} - Season {cwl_data.get('season', 'Unknown')}"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:I1")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Clan info summary
        row = 3
        ws[f"A{row}"] = "Clan Information"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        info_data = [
            ("Clan Name", cwl_data.get('clan_name', 'Unknown')),
            ("Clan Tag", cwl_data.get('clan_tag', 'Unknown')),
            ("Season", cwl_data.get('season', 'Unknown')),
            ("League", get_war_league_label(cwl_data.get('league_id', 0))),
            ("Final Position", cwl_data.get('final_rank', 'Unknown')),
            ("Total Stars", cwl_data.get('total_stars', 0)),
            ("Total Destruction", f"{cwl_data.get('total_destruction', 0):.1f}%"),
            ("Average Stars per Attack", f"{cwl_data.get('average_stars', 0):.2f}"),
        ]

        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        # Add some spacing
        row += 2

        # Members performance table
        ws[f"A{row}"] = "Member Performance"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        # Headers
        headers = [
            "Player Name", "Player Tag", "Town Hall", "Total Attacks", "Total Stars", 
            "Average Stars", "Total Destruction %", "Average Destruction %", "Performance Score"
        ]
        
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i)
            cell.value = header

        start_row = row

        # Add member data
        members_data = cwl_data.get('members', [])
        for member in members_data:
            row += 1
            member_data = [
                member.get('name', 'Unknown'),
                member.get('tag', 'Unknown'),
                member.get('townhall_level', 0),
                member.get('attack_count', 0),
                member.get('total_stars', 0),
                f"{member.get('average_stars', 0):.2f}",
                f"{member.get('total_destruction', 0):.1f}%",
                f"{member.get('average_destruction', 0):.1f}%",
                f"{member.get('performance_score', 0):.2f}"
            ]
            
            for i, value in enumerate(member_data, 1):
                ws.cell(row=row, column=i, value=value)

        end_row = row
        format_table(ws, start_row, end_row, "CWLSummary")

        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            col_letter = column_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(length + 2, 50)

        # Save to temporary file
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)

        # Generate filename
        clean_clan_name = "".join(c for c in cwl_data.get('clan_name', 'Clan') if c.isalnum() or c in (' ', '-', '_')).rstrip()
        clean_tag = clan_tag.replace('#', '')
        filename = f"cwl_summary_{clean_clan_name}_{clean_tag}_season_{cwl_data.get('season', 'unknown')}.xlsx"

        return FileResponse(
            path=tmp.name,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating CWL export: {str(e)}")


@router.post("/war/player-stats", name="Export player war statistics to Excel")
async def export_player_war_stats(filter: PlayerWarhitsFilter):
    # Get war hits data using existing logic
    client = coc.Client(raw_attribute=True)
    START = pend.from_timestamp(filter.timestamp_start, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')
    END = pend.from_timestamp(filter.timestamp_end, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')

    player_tag = fix_tag(filter.player_tags[0])
    pipeline = [
        {"$match": {
            "$and": [
                {"$or": [
                    {"data.clan.members.tag": player_tag},
                    {"data.opponent.members.tag": player_tag}
                ]},
                {"data.preparationStartTime": {"$gte": START}},
                {"data.preparationStartTime": {"$lte": END}}
            ]
        }},
        {"$sort": {"data.preparationStartTime": -1}},
        {"$limit": filter.limit or 10000},
        {"$unset": ["_id"]},
        {"$project": {"data": "$data"}},
    ]

    wars_docs = await mongo.clan_wars.aggregate(pipeline, allowDiskUse=True).to_list(length=None)

    result = await collect_player_hits_from_wars(
        wars_docs,
        tags_to_include=[player_tag],
        clan_tags=None,
        filter=filter,
        client=client
    )

    player_results = result["items"]

    if not player_results:
        raise HTTPException(status_code=404, detail="No war hits found for the specified player and filters")

    # Extract individual attacks from war data
    war_hits = []
    for player in player_results:
        for war_info in player.get("wars", []):
            war_data = war_info.get("war_data", {})
            member_data = war_info.get("members", [{}])[0]

            # Extract attacks from this war
            for attack in member_data.get("attacks", []):
                war_hits.append({
                    "attacker_name": attack.get("attacker", {}).get("name", ""),
                    "attacker_tag": attack.get("attacker", {}).get("tag", ""),
                    "attacker_townhall": attack.get("attacker", {}).get("townhallLevel", ""),
                    "war_date": datetime.strptime(war_data.get("preparationStartTime", ""),
                                                  '%Y%m%dT%H%M%S.%fZ').strftime('%Y-%m-%d %H:%M') if war_data.get(
                        "preparationStartTime") else "",
                    "war_type": "CWL" if war_data.get("season") else "Regular",
                    "attacker_map_position": attack.get("attacker", {}).get("mapPosition", ""),
                    "defender_name": attack.get("defender", {}).get("name", ""),
                    "defender_tag": attack.get("defender", {}).get("tag", ""),
                    "defender_townhall": attack.get("defender", {}).get("townhallLevel", ""),
                    "defender_map_position": attack.get("defender", {}).get("mapPosition", ""),
                    "stars": attack.get("stars", 0),
                    "destruction_percentage": attack.get("destructionPercentage", 0),
                    "order": attack.get("order", ""),
                    "fresh_attack": attack.get("freshAttack", attack.get("fresh", False)),
                    "war_tag": war_data.get("tag", ""),
                    "war_state": war_data.get("state", "")
                })

    if not war_hits:
        raise HTTPException(status_code=404, detail="No individual attacks found for the specified player and filters")

    # Helper function to get filter summary from filter object
    def get_filter_summary():
        filters = []

        # Only add filters that were actually applied (check for None explicitly)
        if filter.season is not None and str(filter.season).strip() and str(filter.season).strip() != 'None':
            filters.append(f"Season {filter.season}")
        if filter.type and filter.type != "all":
            if isinstance(filter.type, list):
                filters.append(f"War type: {', '.join(filter.type)}")
            else:
                filters.append(f"War type: {filter.type}")
        if filter.own_th is not None:
            if isinstance(filter.own_th, list):
                filters.append(f"Own TH: {', '.join(map(str, filter.own_th))}")
            else:
                filters.append(f"Own TH: {filter.own_th}")
        if filter.enemy_th is not None:
            if isinstance(filter.enemy_th, list):
                filters.append(f"Enemy TH: {', '.join(map(str, filter.enemy_th))}")
            else:
                filters.append(f"Enemy TH: {filter.enemy_th}")
        if filter.stars is not None:
            filters.append(f"Stars: {', '.join(map(str, filter.stars))}")
        if filter.min_destruction is not None and filter.min_destruction > 0:
            if filter.max_destruction is not None and filter.max_destruction < 100:
                filters.append(f"Destruction: {filter.min_destruction}%-{filter.max_destruction}%")
            else:
                filters.append(f"Destruction: ≥{filter.min_destruction}%")
        elif filter.max_destruction is not None and filter.max_destruction < 100:
            filters.append(f"Destruction: ≤{filter.max_destruction}%")
        if filter.map_position_min is not None and filter.map_position_min > 1:
            if filter.map_position_max is not None and filter.map_position_max < 50:
                filters.append(f"Map position: {filter.map_position_min}-{filter.map_position_max}")
            else:
                filters.append(f"Map position: ≥{filter.map_position_min}")
        elif filter.map_position_max is not None and filter.map_position_max < 50:
            filters.append(f"Map position: ≤{filter.map_position_max}")
        if filter.fresh_only:
            filters.append("Fresh attacks only")
        # Only show date range if it's not the default values (0 and 2527625513)
        if filter.timestamp_start != 0 or filter.timestamp_end != 2527625513:
            start_date = datetime.fromtimestamp(filter.timestamp_start).strftime('%Y-%m-%d') if filter.timestamp_start > 0 else "Beginning"
            end_date = datetime.fromtimestamp(filter.timestamp_end).strftime('%Y-%m-%d') if filter.timestamp_end < 2527625513 else "Present"
            filters.append(f"Date range: {start_date} to {end_date}")

        return "No filters applied" if not filters else "; ".join(filters)

    # Use the new modular export function
    tmp = await export_player_war_stats_to_excel(
        war_hits=war_hits,
        player_tag=player_tag,
        filter_summary=get_filter_summary()
    )

    # Generate filename and return file
    # Get player name from war hits for filename
    player_name = next((hit.get("attacker_name", "") for hit in war_hits if hit.get("attacker_name")), "Player")
    
    # Clean player name for filename
    clean_player_name = "".join(c for c in player_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    clean_tag = player_tag.replace('#', '')
    filename = f"war_stats_{clean_player_name}_{clean_tag}.xlsx"

    return FileResponse(
        path=tmp.name,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )