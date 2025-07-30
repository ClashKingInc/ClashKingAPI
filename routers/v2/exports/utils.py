"""
Excel export utilities for ClashKing API
"""
import aiohttp
import tempfile
from typing import Optional, List
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell


# ClashKing theme colors and styles
CLASHKING_THEME = {
    "header_fill_red": PatternFill(start_color="D00000", end_color="D00000", fill_type="solid"),
    "header_fill_black": PatternFill(start_color="000000", end_color="000000", fill_type="solid"),
    "header_font_white": Font(color="FFFFFF", bold=True),
    "data_font_black": Font(color="000000"),
    "title_fill": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
    "data_fill_white": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
}

# Standard ClashKing logo URL
CLASHKING_LOGO_URL = "https://assets.clashk.ing/logos/crown-text-white-bg/BqlEp974170917vB1qK0zunfANJCGi0W031dTksEq7KQ9LoXWMFk0u77unHJa.png"


async def insert_logo_from_cdn(sheet: Worksheet, image_url: str, anchor_cell="A1", height=80):
    """Insert a logo from CDN URL into an Excel sheet"""
    async with aiohttp.ClientSession() as session:
        async with session.get(image_url) as resp:
            if resp.status != 200:
                raise Exception(f"Failed to download logo from CDN: {resp.status}")
            image_data = await resp.read()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
        tmp_img.write(image_data)
        tmp_img_path = tmp_img.name

    logo = OpenpyxlImage(tmp_img_path)
    logo.height = height
    logo.width = int(height * 3)
    sheet.add_image(logo, anchor_cell)


async def add_clashking_logo_to_sheet(sheet: Worksheet, anchor_cell="H1", height=50):
    """Add the standard ClashKing logo to a sheet"""
    await insert_logo_from_cdn(
        sheet,
        image_url=CLASHKING_LOGO_URL,
        anchor_cell=anchor_cell,
        height=height
    )


def format_table(sheet: Worksheet, start_row: int, end_row: int, table_name_hint: str, highlight_rows: Optional[List[int]] = None):
    """
    Format an Excel table with ClashKing styling
    
    Args:
        sheet: The worksheet containing the table
        start_row: Row number where the table headers start
        end_row: Row number where the table data ends
        table_name_hint: Hint for table naming (not used in current implementation)
        highlight_rows: List of row numbers to highlight with gold color
    """
    # Simplified table formatting without Excel Table objects to avoid corruption
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    # Style header row
    for cell in sheet[start_row]:
        if cell.value is None:
            cell.value = ""
        else:
            cell.value = str(cell.value)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = CLASHKING_THEME["header_fill_red"]
        cell.font = CLASHKING_THEME["header_font_white"]
        cell.border = thin_border

    # Style data rows with alternating colors and highlights
    for row_num, row in enumerate(sheet.iter_rows(min_row=start_row + 1, max_row=end_row)):
        actual_row_num = start_row + 1 + row_num
        
        # Check if this row should be highlighted (current player TH)
        should_highlight = highlight_rows and actual_row_num in highlight_rows
        
        if should_highlight:
            # Highlight current player TH with gold/yellow color
            row_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
        else:
            # Alternate row colors for readability
            row_fill = CLASHKING_THEME["data_fill_white"] if row_num % 2 == 0 else PatternFill(
                start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
            )

        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.fill = row_fill
            cell.border = thin_border

    # Auto-adjust column widths
    for column_cells in sheet.columns:
        first_real_cell = next((cell for cell in column_cells if isinstance(cell, Cell)), None)
        if not first_real_cell:
            continue
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = first_real_cell.column_letter
        sheet.column_dimensions[col_letter].width = min(length + 2, 50)  # Cap at 50 chars width


def add_title_to_sheet(sheet: Worksheet, title: str, merge_range: str = "A1:P1"):
    """Add a formatted title to a sheet"""
    row = sheet.max_row + 3
    sheet.merge_cells(merge_range.replace("1", str(row)))
    cell = sheet.cell(row=row, column=1)
    cell.value = title
    cell.alignment = Alignment(horizontal="center")
    cell.fill = CLASHKING_THEME["title_fill"]
    cell.font = Font(bold=True, size=14)


def add_filter_summary_to_sheet(sheet: Worksheet, filter_summary: str, merge_range: str = "A1:P1"):
    """Add filter summary under title"""
    row = sheet.max_row + 1
    sheet.merge_cells(merge_range.replace("1", str(row)))
    cell = sheet.cell(row=row, column=1)
    cell.value = filter_summary
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(italic=True, size=10)
    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")


def add_table_headers(sheet: Worksheet, headers: List[str]):
    """Add table headers to a sheet"""
    row = sheet.max_row + 1
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=i)
        cell.value = header
    return row


def add_summary_data(sheet: Worksheet, data: List[tuple], preserve_numeric_types: bool = True):
    """
    Add summary data to a sheet with proper type preservation
    
    Args:
        sheet: The worksheet
        data: List of (label, value) tuples
        preserve_numeric_types: Whether to keep numeric values as numbers
    """
    for stat_name, stat_value in data:
        if preserve_numeric_types and isinstance(stat_value, (int, float)):
            sheet.append([str(stat_name), stat_value])
        else:
            sheet.append([str(stat_name), str(stat_value)])


def add_insights_section(sheet: Worksheet, insights: List[str], merge_columns: str = "A:I"):
    """Add properly formatted insights as merged cells"""
    for insight in insights:
        row = sheet.max_row + 1
        merge_range = f"{merge_columns[0]}{row}:{merge_columns[-1]}{row}"
        sheet.merge_cells(merge_range)
        cell = sheet.cell(row=row, column=1)
        cell.value = insight
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(italic=True, size=11)
        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Light blue background


def add_section_title(sheet: Worksheet, title: str, merge_columns: str = "A:I"):
    """Add a section title with proper formatting"""
    row = sheet.max_row + 3
    merge_range = f"{merge_columns[0]}{row}:{merge_columns[-1]}{row}"
    sheet.merge_cells(merge_range)
    cell = sheet.cell(row=row, column=1)
    cell.value = title
    cell.alignment = Alignment(horizontal="center")
    cell.fill = CLASHKING_THEME["header_fill_black"]
    cell.font = CLASHKING_THEME["header_font_white"]


def add_empty_row(sheet: Worksheet):
    """Add an empty row for spacing"""
    sheet.append([])