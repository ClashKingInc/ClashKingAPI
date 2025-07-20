import asyncio
import json
import tempfile
import uuid
import coc
import sentry_sdk
from datetime import datetime

import aiohttp
from fastapi.responses import JSONResponse
import pendulum as pend
from fastapi import HTTPException
from fastapi import APIRouter, Request
from openpyxl.cell import Cell

from routers.v2.clan.models import ClanTagsRequest
from routers.v2.war.models import PlayerWarhitsFilter, ClanWarHitsFilter
from routers.v2.war.utils import fetch_current_war_info_bypass, fetch_league_info, ranking_create, \
    fetch_war_league_infos, enrich_league_info, collect_player_hits_from_wars
from utils.time import is_cwl
from utils.utils import fix_tag, remove_id_fields
from utils.database import MongoClient as mongo

from fastapi.responses import FileResponse
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as OpenpyxlImage

router = APIRouter(prefix="/v2", tags=["War"], include_in_schema=True)


@router.get("/war/{clan_tag}/previous",
            name="Previous Wars for a clan")
async def war_previous(
        clan_tag: str,
        request: Request = Request,
        timestamp_start: int = 0,
        timestamp_end: int = 9999999999,
        include_cwl: bool = False,
        limit: int = 50,
):
    clan_tag = fix_tag(clan_tag)
    START = pend.from_timestamp(timestamp_start, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')
    END = pend.from_timestamp(timestamp_end, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')

    query = {
        "$and": [
            {"$or": [{"data.clan.tag": clan_tag}, {"data.opponent.tag": clan_tag}]},
            {"data.preparationStartTime": {"$gte": START}},
            {"data.preparationStartTime": {"$lte": END}}
        ]
    }

    if not include_cwl:
        query["$and"].append({"data.season": None})

    full_wars = await mongo.clan_wars.find(query).sort("data.endTime", -1).limit(limit).to_list(length=None)

    # early on we had some duplicate wars, so just filter them out
    found_ids = set()
    new_wars = []
    for war in full_wars:
        id = war.get("data").get("preparationStartTime")
        if id in found_ids:
            continue
        war.pop("_response_retry", None)
        new_wars.append(war.get("data"))
        found_ids.add(id)

    return remove_id_fields({"items": new_wars})


@router.get("/cwl/{clan_tag}/ranking-history", name="CWL ranking history for a clan")
async def cwl_ranking_history(clan_tag: str, request: Request):
    clan_tag = fix_tag(clan_tag)

    # Fetch all CWL group documents containing the clan
    results = await mongo.cwl_groups.find({"data.clans.tag": clan_tag}, {"data.clans": 0}).to_list(length=None)
    if not results:
        raise HTTPException(status_code=404, detail="No CWL Data Found")

    # Get league changes for the clan
    clan_data = await mongo.basic_clan.find_one({"tag": clan_tag}, {"changes.clanWarLeague": 1})
    cwl_changes = clan_data.get("changes", {}).get("clanWarLeague", {})

    # Collect all war tags across all CWL results (across all seasons)
    all_war_tags = set()
    for cwl_result in results:
        rounds = cwl_result["data"].get("rounds", [])
        for rnd in rounds:
            for tag in rnd.get("warTags", []):
                if tag:
                    all_war_tags.add(tag)

    # Fetch all war documents in one go using the union of war tags
    if all_war_tags:
        matching_wars_data = await mongo.clan_wars.find({
            "data.tag": {"$in": list(all_war_tags)}
        },
            {"data.clan.members": 0, "data.opponent.members": 0}
        ).to_list(length=None)
        # Build a lookup dictionary keyed by war tag
        war_lookup = {w["data"]["tag"]: w["data"] for w in matching_wars_data}
    else:
        war_lookup = {}

    ranking_results = []
    for cwl_result in results:
        season = cwl_result["data"].get("season")
        rounds = cwl_result["data"].get("rounds", [])

        # Replace each war tag with the corresponding war data,
        # but only include wars that match the current season.
        for rnd in rounds:
            rnd["warTags"] = [
                war_lookup.get(tag)
                for tag in rnd.get("warTags", [])
                if war_lookup.get(tag) and war_lookup.get(tag).get("season") == season
            ]

        cwl_data = cwl_result["data"]
        cwl_data["rounds"] = rounds
        ranking = ranking_create(data=cwl_data)

        # Get the ranking for our clan tag along with its index (place)
        ranking_data = next(
            (
                {'rank': idx, **item}
                for idx, item in enumerate(ranking, start=1)
                if item["tag"] == clan_tag
            ),
            None
        )
        if ranking_data is None:
            continue

        # Calculate season offset and check league changes
        season_offset = pend.date(
            year=int(season[:4]),
            month=int(season[-2:]),
            day=1
        ).subtract(months=1).strftime('%Y-%m')
        if season_offset not in cwl_changes:
            continue

        league = cwl_changes[season_offset].get("league")
        ranking_results.append({"season": season, "league": league, **ranking_data})

    return {"items": sorted(ranking_results, key=lambda x: x["season"], reverse=True)}


@router.get("/cwl/league-thresholds", name="Promo and demotion thresholds for CWL leagues")
async def cwl_league_thresholds(request: Request):
    return {
        "items": [
            {
                "id": 48000001,
                "name": "Bronze League III",
                "promo": 3,
                "demote": 9
            },
            {
                "id": 48000002,
                "name": "Bronze League II",
                "promo": 3,
                "demote": 8
            },
            {
                "id": 48000003,
                "name": "Bronze League I",
                "promo": 3,
                "demote": 8
            },
            {
                "id": 48000004,
                "name": "Silver League III",
                "promo": 2,
                "demote": 8
            },
            {
                "id": 48000005,
                "name": "Silver League II",
                "promo": 2,
                "demote": 7
            },
            {
                "id": 48000006,
                "name": "Silver League I",
                "promo": 2,
                "demote": 7
            },
            {
                "id": 48000007,
                "name": "Gold League III",
                "promo": 2,
                "demote": 7
            },
            {
                "id": 48000008,
                "name": "Gold League II",
                "promo": 2,
                "demote": 7
            },
            {
                "id": 48000009,
                "name": "Gold League I",
                "promo": 2,
                "demote": 7
            },
            {
                "id": 48000010,
                "name": "Crystal League III",
                "promo": 2,
                "demote": 7
            },
            {
                "id": 48000011,
                "name": "Crystal League II",
                "promo": 2,
                "demote": 7
            },
            {
                "id": 48000012,
                "name": "Crystal League I",
                "promo": 1,
                "demote": 7
            },
            {
                "id": 48000013,
                "name": "Master League III",
                "promo": 1,
                "demote": 7
            },
            {
                "id": 48000014,
                "name": "Master League II",
                "promo": 1,
                "demote": 7
            },
            {
                "id": 48000015,
                "name": "Master League I",
                "promo": 1,
                "demote": 7
            },
            {
                "id": 48000016,
                "name": "Champion League III",
                "promo": 1,
                "demote": 7
            },
            {
                "id": 48000017,
                "name": "Champion League II",
                "promo": 1,
                "demote": 7
            },
            {
                "id": 48000018,
                "name": "Champion League I",
                "promo": 0,
                "demote": 6
            }
        ]
    }


@router.get(
    "/war/{clan_tag}/war-summary",
    name="Get full war and CWL summary for a clan, including war state, CWL rounds and war details"
)
async def get_clan_war_summary(clan_tag: str):
    async with aiohttp.ClientSession() as session:
        war_info = await fetch_current_war_info_bypass(clan_tag, session)
        league_info = None
        war_league_infos = []

        if is_cwl():
            league_info = await fetch_league_info(clan_tag, session)
            if league_info and "rounds" in league_info:
                for round_entry in league_info["rounds"]:
                    war_tags = round_entry.get("warTags", [])
                    war_league_infos.extend(await fetch_war_league_infos(war_tags, session))

                league_info = await enrich_league_info(league_info, war_league_infos, session)

        return JSONResponse(content={
            "isInWar": war_info and war_info.get("state") == "war",
            "isInCwl": league_info is not None and war_info and war_info.get("state") == "notInWar",
            "war_info": war_info,
            "league_info": league_info,
            "war_league_infos": war_league_infos
        })


@router.post("/war/war-summary", name="Get full war and CWL summary for multiple clans")
async def get_multiple_clan_war_summary(body: ClanTagsRequest, request: Request):
    if not body.clan_tags:
        raise HTTPException(status_code=400, detail="clan_tags cannot be empty")

    async with aiohttp.ClientSession() as session:

        async def process_clan(clan_tag: str):
            war_info = await fetch_current_war_info_bypass(clan_tag, session)
            league_info = None
            war_league_infos = []

            if is_cwl():
                league_info = await fetch_league_info(clan_tag, session)
                if league_info and "rounds" in league_info:
                    war_tags = [tag for r in league_info["rounds"] for tag in r.get("warTags", [])]
                    war_league_infos = await fetch_war_league_infos(war_tags, session)
                    league_info = await enrich_league_info(league_info, war_league_infos, session)

            return {
                "clan_tag": clan_tag,
                "isInWar": war_info and war_info.get("state") == "war",
                "isInCwl": league_info is not None and war_info and war_info.get("state") == "notInWar",
                "war_info": war_info,
                "league_info": league_info,
                "war_league_infos": war_league_infos
            }

        results = await asyncio.gather(*(process_clan(tag) for tag in body.clan_tags))
        return JSONResponse(content={"items": results})


@router.get("/war/cwl-summary/export", name="Export CWL summary and members stats to Excel")
async def export_cwl_summary_to_excel(tag: str):
    def format_table(sheet, start_row, end_row, table_name_hint):
        table_name = f"{table_name_hint}_{uuid.uuid4().hex[:8]}"[:31]

        thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )

        for cell in sheet[start_row]:
            cell.alignment = Alignment(horizontal="center")
            cell.fill = clashking_theme["header_fill_red"]
            cell.font = clashking_theme["header_font_white"]
            cell.border = thin_border

        for row in sheet.iter_rows(min_row=start_row + 1, max_row=end_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.fill = clashking_theme["data_fill_white"]
                cell.border = thin_border

        for column_cells in sheet.columns:
            first_real_cell = next((cell for cell in column_cells if isinstance(cell, Cell)), None)
            if not first_real_cell:
                continue
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = first_real_cell.column_letter
            sheet.column_dimensions[col_letter].width = length + 2

        last_col_letter = sheet[end_row][len(sheet[end_row]) - 1].column_letter
        table_range = f"A{start_row}:{last_col_letter}{end_row}"
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

    async def insert_logo_from_cdn(sheet, image_url: str, anchor_cell="A1", height=80):
        # Download the image to a temporary file
        async with aiohttp.ClientSession() as session:
            async with session.get(image_url) as resp:
                if resp.status != 200:
                    raise Exception(f"Failed to download logo from CDN: {resp.status}")
                image_data = await resp.read()

        # Write image to a temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
            tmp_img.write(image_data)
            tmp_img_path = tmp_img.name

        # Insert into Excel
        logo = OpenpyxlImage(tmp_img_path)
        logo.height = height
        logo.width = int(height * 3)
        sheet.add_image(logo, anchor_cell)

    clashking_theme = {
        "header_fill_red": PatternFill(start_color="D00000", end_color="D00000", fill_type="solid"),
        "header_fill_black": PatternFill(start_color="000000", end_color="000000", fill_type="solid"),
        "header_font_white": Font(color="FFFFFF", bold=True),
        "data_font_black": Font(color="000000"),
        "title_fill": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
        "data_fill_white": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    }

    summary = await get_clan_war_summary(tag)
    if isinstance(summary, JSONResponse):
        summary = json.loads(summary.body)

    league_info = summary.get("league_info")

    if not league_info:
        raise HTTPException(status_code=404, detail="No league info available")

    wb = Workbook()
    ws_clan = wb.active
    ws_clan.title = "Summary"

    clans = sorted(league_info.get("clans", []), key=lambda c: c.get("rank", 0))

    await insert_logo_from_cdn(
        ws_clan,
        image_url="https://assets.clashk.ing/logos/crown-text-white-bg/BqlEp974170917vB1qK0zunfANJCGi0W031dTksEq7KQ9LoXWMFk0u77unHJa.png",
        anchor_cell="F1",
        height=50
    )

    row = ws_clan.max_row + 3
    ws_clan.merge_cells(f"A{row}:J{row}")
    cell = ws_clan.cell(row=row, column=1)
    cell.value = f"[{league_info.get('season', '')}] 🏆 {league_info.get('war_league', '')}"
    cell.alignment = Alignment(horizontal="center")
    cell.fill = clashking_theme["title_fill"]
    cell.font = Font(bold=True, size=14)

    row = ws_clan.max_row + 1
    ws_clan.merge_cells(f"A{row}:J{row}")
    cell = ws_clan.cell(row=row, column=1)
    cell.value = "CWL Clan Rankings"
    cell.alignment = Alignment(horizontal="center")
    cell.fill = clashking_theme["header_fill_black"]
    cell.font = clashking_theme["header_font_white"]

    ws_clan.append([
        "Rank", "Name", "Tag", "Level", "Wars Played",
        "Total Stars", "Total Destruction %", "Destruction Taken %",
        "Attack Count", "Missed Attacks"
    ])
    start_row = ws_clan.max_row

    for clan in clans:
        ws_clan.append([
            clan.get("rank"),
            clan.get("name"),
            clan.get("tag"),
            clan.get("clanLevel"),
            clan.get("wars_played"),
            clan.get("total_stars"),
            clan.get("total_destruction"),
            clan.get("total_destruction_inflicted"),
            clan.get("attack_count"),
            clan.get("missed_attacks"),
        ])
    end_row = ws_clan.max_row
    format_table(ws_clan, start_row, end_row, "ClanSummary")

    if tag is None:
        sentry_sdk.capture_message("clan tag is None in export_cwl_summary_to_excel", level="error")
        raise HTTPException(status_code=400, detail="clan tag cannot be None")
    tag = tag.replace("!", "#")
    for clan in clans:
        sheet = wb.create_sheet(title=clan.get("name", "Clan")[:30])
        await insert_logo_from_cdn(
            sheet,
            image_url="https://assets.clashk.ing/logos/crown-text-white-bg/BqlEp974170917vB1qK0zunfANJCGi0W031dTksEq7KQ9LoXWMFk0u77unHJa.png",
            anchor_cell="N1",
            height=50
        )

        row = sheet.max_row + 3
        sheet.merge_cells(f"A{row}:V{row}")
        cell = sheet.cell(row=row, column=1)
        cell.value = "⚔️ Attacks"
        cell.alignment = Alignment(horizontal="center")
        cell.fill = clashking_theme["header_fill_black"]
        cell.font = clashking_theme["header_font_white"]

        headers_attack = [
            "Name", "Tag", "TH", "War Participated", "Attacks Done", "Missed",
            "Stars", "Avg Stars", "3 Stars",
            "2 Stars", "1 Star", "0 Star", "3 Stars %", "0-1 Stars %",
            "Total Destruction", "Avg Destruction", "Avg Map Position", "Avg Opponent Map Position",
            "Avg Order", "Avg Opponent TH Level", "Lower TH", "Upper TH"
        ]
        sheet.append(headers_attack)
        attack_start_row = sheet.max_row

        for member in clan.get("members", []):
            if member.get("avgMapPosition"):
                a = member.get("attacks", {}) or {}
                w = a.get("missed_attacks", 0) + a.get("attack_count", 0)
                s = a.get("stars", 0)
                d = a.get("total_destruction", 0)

                sheet.append([
                    member.get("name"),
                    member.get("tag"),
                    member.get("townHallLevel"),
                    w,
                    a.get("attack_count", 0),
                    a.get("missed_attacks", 0),
                    s,
                    round(s / w if w > 0 else 0, 2),
                    sum((a.get("3_stars") or {}).values()),
                    sum((a.get("2_stars") or {}).values()),
                    sum((a.get("1_star") or {}).values()),
                    sum((a.get("0_star") or {}).values()),
                    round((sum((a.get("3_stars") or {}).values()) / w) * 100 if w > 0 else 0, 2),
                    round(((sum((a.get("0_star") or {}).values()) + sum(
                        (a.get("1_star") or {}).values())) * 100 / w) if w > 0 else 0, 2),
                    d,
                    round(d / w if w > 0 else 0, 2),
                    member.get("avgMapPosition"),
                    member.get("avgOpponentPosition"),
                    member.get("avgAttackOrder"),
                    member.get("avgOpponentTownHallLevel"),
                    member.get("attackLowerTHLevel"),
                    member.get("attackUpperTHLevel"),
                ])

        attack_end_row = sheet.max_row
        format_table(sheet, attack_start_row, attack_end_row, "AttacksTable")

        sheet.append([])  # Empty row
        sheet.append([])  # Empty row

        row = sheet.max_row + 2
        sheet.merge_cells(f"A{row}:V{row}")
        cell = sheet.cell(row=row, column=1)
        cell.value = "🛡️ Defenses"
        cell.alignment = Alignment(horizontal="center")
        cell.fill = clashking_theme["header_fill_black"]
        cell.font = clashking_theme["header_font_white"]

        headers_defense = [
            "Name", "Tag", "TH", "War Participated", "Defenses Received", "Missed",
            "Stars Taken", "Avg Stars", "3 Stars",
            "2 Stars", "1 Star", "0 Star", "3 Stars %", "0-1 Stars %",
            "Total Destruction", "Avg Destruction", "Avg Attacker Map Position", "Avg Map Position",
            "Avg Opponent Order",
            "Avg Opponent TH Level", "Lower TH", "Upper TH"
        ]
        sheet.append(headers_defense)
        defense_start_row = sheet.max_row

        for member in clan.get("members", []):
            if member.get("avgMapPosition"):
                attacks = member.get("attacks", {}) or {}
                war_participated = attacks.get("missed_attacks", 0) + attacks.get("attack_count", 0)
                defenses = member.get("defense", {}) or {}
                defense_count = defenses.get("defense_count", 0)
                stars_total = defenses.get("stars", 0)
                destruction_total = defenses.get("total_destruction", 0)
                missed_defenses = defenses.get("missed_defenses", 0)

                three_stars = sum((defenses.get("3_stars") or {}).values())
                two_stars = sum((defenses.get("2_stars") or {}).values())
                one_star = sum((defenses.get("1_star") or {}).values())
                zero_star = sum((defenses.get("0_star") or {}).values())

                sheet.append([
                    member.get("name"),
                    member.get("tag"),
                    member.get("townHallLevel"),
                    war_participated,
                    defense_count,
                    missed_defenses,
                    stars_total,
                    round(stars_total / war_participated * 100 if war_participated > 0 else 0, 2),
                    three_stars,
                    two_stars,
                    one_star,
                    zero_star,
                    round((three_stars / war_participated) * 100 if war_participated > 0 else 0, 2),
                    round(((zero_star + one_star) * 100 / war_participated) if war_participated > 0 else 0, 2),
                    destruction_total,
                    round(destruction_total / war_participated if war_participated > 0 else 0, 2),
                    member.get("avgOpponentPosition"),
                    member.get("avgMapPosition"),
                    member.get("avgDefenseOrder"),
                    member.get("avgAttackerTownHallLevel"),
                    member.get("defenseLowerTHLevel"),
                    member.get("defenseUpperTHLevel"),
                ])

        defense_end_row = sheet.max_row
        format_table(sheet, defense_start_row, defense_end_row, "DefensesTable")

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.seek(0)

    clan_name = next((c for c in clans if c.get("tag") == tag), None).get("name", "clan")
    season = league_info.get("season")
    filename = f"cwl_summary_{clan_name}_{season}.xlsx"
    return FileResponse(
        path=tmp.name,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.post("/war/players/warhits")
async def players_warhits_stats(filter: PlayerWarhitsFilter, request: Request):
    client = coc.Client(raw_attribute=True)
    START = pend.from_timestamp(filter.timestamp_start, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')
    END = pend.from_timestamp(filter.timestamp_end, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')

    async def fetch_player(tag: str):
        player_tag = fix_tag(tag)
        pipeline = [
            {"$match": {
                "$and": [
                    {"$or": [
                        {"data.clan.members.tag": player_tag},
                        {"data.opponent.members.tag": player_tag}
                    ]},
                    {"data.preparationStartTime": {"$gte": START}},
                    {"data.preparationStartTime": {"$lte": END}}
                ]
            }},
            {"$sort": {"data.preparationStartTime": -1}},
            {"$limit": filter.limit or 50},
            {"$unset": ["_id"]},
            {"$project": {"data": "$data"}},
        ]

        wars_docs = await mongo.clan_wars.aggregate(pipeline, allowDiskUse=True).to_list(length=None)

        result = await collect_player_hits_from_wars(
            wars_docs,
            tags_to_include=[player_tag],
            clan_tags=None,
            filter=filter,
            client=client
        )
        return result["items"]

    player_tasks = [fetch_player(tag) for tag in filter.player_tags]
    results_per_player = await asyncio.gather(*player_tasks)
    results = [item for sublist in results_per_player for item in sublist]  # flatten

    return {"items": results}


@router.get("/war/players/warhits/export", name="Export player war statistics to Excel")
async def export_player_war_stats(
        tag: str,
        timestamp_start: int = 0,
        timestamp_end: int = 9999999999,
        season: str = None,
        war_type: str = None,
        own_th: str = None,
        enemy_th: str = None,
        stars: str = None,
        min_destruction: int = None,
        max_destruction: int = None,
        map_position_min: int = None,
        map_position_max: int = None,
        fresh_only: bool = False,
        limit: int = 500
):
    def format_table(sheet, start_row, end_row, table_name_hint, highlight_rows=None):
        # Simplified table formatting without Excel Table objects to avoid corruption
        thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )

        # Style header row
        for cell in sheet[start_row]:
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = str(cell.value)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = clashking_theme["header_fill_red"]
            cell.font = clashking_theme["header_font_white"]
            cell.border = thin_border

        # Style data rows with alternating colors and highlights
        for row_num, row in enumerate(sheet.iter_rows(min_row=start_row + 1, max_row=end_row)):
            actual_row_num = start_row + 1 + row_num
            
            # Check if this row should be highlighted (current player TH)
            should_highlight = highlight_rows and actual_row_num in highlight_rows
            
            if should_highlight:
                # Highlight current player TH with gold/yellow color
                row_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
            else:
                # Alternate row colors for readability
                row_fill = clashking_theme["data_fill_white"] if row_num % 2 == 0 else PatternFill(start_color="F8F8F8",
                                                                                                   end_color="F8F8F8",
                                                                                                   fill_type="solid")

            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.fill = row_fill
                cell.border = thin_border

        # Auto-adjust column widths
        for column_cells in sheet.columns:
            first_real_cell = next((cell for cell in column_cells if isinstance(cell, Cell)), None)
            if not first_real_cell:
                continue
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = first_real_cell.column_letter
            sheet.column_dimensions[col_letter].width = min(length + 2, 50)  # Cap at 50 chars width

    async def insert_logo_from_cdn(sheet, image_url: str, anchor_cell="A1", height=80):
        async with aiohttp.ClientSession() as session:
            async with session.get(image_url) as resp:
                if resp.status != 200:
                    raise Exception(f"Failed to download logo from CDN: {resp.status}")
                image_data = await resp.read()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
            tmp_img.write(image_data)
            tmp_img_path = tmp_img.name

        logo = OpenpyxlImage(tmp_img_path)
        logo.height = height
        logo.width = int(height * 3)
        sheet.add_image(logo, anchor_cell)

    clashking_theme = {
        "header_fill_red": PatternFill(start_color="D00000", end_color="D00000", fill_type="solid"),
        "header_fill_black": PatternFill(start_color="000000", end_color="000000", fill_type="solid"),
        "header_font_white": Font(color="FFFFFF", bold=True),
        "data_font_black": Font(color="000000"),
        "title_fill": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
        "data_fill_white": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    }

    # Create filter from parameters
    from routers.v2.war.models import PlayerWarhitsFilter
    filter_data = {
        "player_tags": [fix_tag(tag)],
        "timestamp_start": timestamp_start,
        "timestamp_end": timestamp_end,
        "limit": limit
    }

    if season:
        filter_data["season"] = season
    if war_type:
        filter_data["type"] = [war_type]
    if own_th:
        filter_data["own_th"] = [int(x) for x in own_th.split(',')]
    if enemy_th:
        filter_data["enemy_th"] = [int(x) for x in enemy_th.split(',')]
    if stars:
        filter_data["stars"] = [int(x) for x in stars.split(',')]
    if min_destruction is not None:
        filter_data["min_destruction"] = min_destruction
    if max_destruction is not None:
        filter_data["max_destruction"] = max_destruction
    if map_position_min is not None:
        filter_data["map_position_min"] = map_position_min
    if map_position_max is not None:
        filter_data["map_position_max"] = map_position_max
    if fresh_only:
        filter_data["fresh_only"] = fresh_only

    filter = PlayerWarhitsFilter(**filter_data)

    # Get war hits data using existing logic
    client = coc.Client(raw_attribute=True)
    START = pend.from_timestamp(filter.timestamp_start, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')
    END = pend.from_timestamp(filter.timestamp_end, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')

    player_tag = fix_tag(tag)
    pipeline = [
        {"$match": {
            "$and": [
                {"$or": [
                    {"data.clan.members.tag": player_tag},
                    {"data.opponent.members.tag": player_tag}
                ]},
                {"data.preparationStartTime": {"$gte": START}},
                {"data.preparationStartTime": {"$lte": END}}
            ]
        }},
        {"$sort": {"data.preparationStartTime": -1}},
        {"$limit": limit},
        {"$unset": ["_id"]},
        {"$project": {"data": "$data"}},
    ]

    wars_docs = await mongo.clan_wars.aggregate(pipeline, allowDiskUse=True).to_list(length=None)

    result = await collect_player_hits_from_wars(
        wars_docs,
        tags_to_include=[player_tag],
        clan_tags=None,
        filter=filter,
        client=client
    )

    player_results = result["items"]

    if not player_results:
        raise HTTPException(status_code=404, detail="No war hits found for the specified player and filters")

    # Extract individual attacks from war data
    war_hits = []
    for player in player_results:
        for war_info in player.get("wars", []):
            war_data = war_info.get("war_data", {})
            member_data = war_info.get("members", [{}])[0]

            # Extract attacks from this war
            for attack in member_data.get("attacks", []):
                war_hits.append({
                    "attacker_name": attack.get("attacker", {}).get("name", ""),
                    "attacker_tag": attack.get("attacker", {}).get("tag", ""),
                    "attacker_townhall": attack.get("attacker", {}).get("townhallLevel", ""),
                    "war_date": datetime.strptime(war_data.get("preparationStartTime", ""),
                                                  '%Y%m%dT%H%M%S.%fZ').strftime('%Y-%m-%d %H:%M') if war_data.get(
                        "preparationStartTime") else "",
                    "war_type": "CWL" if war_data.get("season") else "Regular",
                    "attacker_map_position": attack.get("attacker", {}).get("mapPosition", ""),
                    "defender_name": attack.get("defender", {}).get("name", ""),
                    "defender_tag": attack.get("defender", {}).get("tag", ""),
                    "defender_townhall": attack.get("defender", {}).get("townhallLevel", ""),
                    "defender_map_position": attack.get("defender", {}).get("mapPosition", ""),
                    "stars": attack.get("stars", 0),
                    "destruction_percentage": attack.get("destructionPercentage", 0),
                    "order": attack.get("order", ""),
                    "fresh_attack": attack.get("freshAttack", attack.get("fresh", False)),
                    "war_tag": war_data.get("tag", ""),
                    "war_state": war_data.get("state", "")
                })

    if not war_hits:
        raise HTTPException(status_code=404, detail="No individual attacks found for the specified player and filters")

    # Store original URL parameters to avoid contamination
    original_params = {
        'season': season,
        'war_type': war_type,
        'own_th': own_th,
        'enemy_th': enemy_th,
        'stars': stars,
        'min_destruction': min_destruction,
        'max_destruction': max_destruction,
        'map_position_min': map_position_min,
        'map_position_max': map_position_max,
        'fresh_only': fresh_only,
        'timestamp_start': timestamp_start,
        'timestamp_end': timestamp_end
    }

    # Helper function to get filter summary
    def get_filter_summary():
        filters = []

        # Use original URL parameters only
        p = original_params

        # Only add filters that were actually applied (check for None explicitly)
        if p['season'] is not None and str(p['season']).strip() and str(p['season']).strip() != 'None':
            filters.append(f"Season {p['season']}")
        if p['war_type'] is not None and str(p['war_type']).strip() and str(p['war_type']).lower() not in ['all',
                                                                                                           'none']:
            filters.append(f"War type: {p['war_type']}")
        if p['own_th'] is not None and str(p['own_th']).strip() and str(p['own_th']).strip() != 'None':
            filters.append(f"Own TH: {p['own_th']}")
        if p['enemy_th'] is not None and str(p['enemy_th']).strip() and str(p['enemy_th']).strip() != 'None':
            filters.append(f"Enemy TH: {p['enemy_th']}")
        if p['stars'] is not None and str(p['stars']).strip() and str(p['stars']).strip() != 'None':
            filters.append(f"Stars: {p['stars']}")
        if p['min_destruction'] is not None and p['min_destruction'] > 0:
            if p['max_destruction'] is not None and p['max_destruction'] < 100:
                filters.append(f"Destruction: {p['min_destruction']}%-{p['max_destruction']}%")
            else:
                filters.append(f"Destruction: ≥{p['min_destruction']}%")
        elif p['max_destruction'] is not None and p['max_destruction'] < 100:
            filters.append(f"Destruction: ≤{p['max_destruction']}%")
        if p['map_position_min'] is not None and p['map_position_min'] > 1:
            if p['map_position_max'] is not None and p['map_position_max'] < 50:
                filters.append(f"Map position: {p['map_position_min']}-{p['map_position_max']}")
            else:
                filters.append(f"Map position: ≥{p['map_position_min']}")
        elif p['map_position_max'] is not None and p['map_position_max'] < 50:
            filters.append(f"Map position: ≤{p['map_position_max']}")
        if p['fresh_only']:
            filters.append("Fresh attacks only")
        # Only show date range if it's not the default full range
        if p['timestamp_start'] > 0 or p['timestamp_end'] < 9999999999:
            start_date = datetime.fromtimestamp(p['timestamp_start']).strftime('%Y-%m-%d') if p[
                                                                                                  'timestamp_start'] > 0 else "Beginning"
            end_date = datetime.fromtimestamp(p['timestamp_end']).strftime('%Y-%m-%d') if p[
                                                                                              'timestamp_end'] < 9999999999 else "Present"
            filters.append(f"Date range: {start_date} to {end_date}")

        return "No filters applied" if not filters else "; ".join(filters)

    # Create Excel workbook
    wb = Workbook()
    
    # Create Summary sheet first (will be the active/first sheet)
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Create Attack Details sheet
    ws_attacks = wb.create_sheet(title="Attack Details")
    
    # Create Defense Details sheet
    ws_defenses = wb.create_sheet(title="Defense Details")

    # Add centered logo to Attack Details sheet
    await insert_logo_from_cdn(
        ws_attacks,
        image_url="https://assets.clashk.ing/logos/crown-text-white-bg/BqlEp974170917vB1qK0zunfANJCGi0W031dTksEq7KQ9LoXWMFk0u77unHJa.png",
        anchor_cell="H1",
        height=50
    )

    # Title for Attack Details
    row = ws_attacks.max_row + 3
    ws_attacks.merge_cells(f"A{row}:P{row}")
    cell = ws_attacks.cell(row=row, column=1)
    cell.value = f"Attack Details for {tag}"
    cell.alignment = Alignment(horizontal="center")
    cell.fill = clashking_theme["title_fill"]
    cell.font = Font(bold=True, size=14)

    # Add filters under title
    row = ws_attacks.max_row + 1
    ws_attacks.merge_cells(f"A{row}:P{row}")
    cell = ws_attacks.cell(row=row, column=1)
    cell.value = get_filter_summary()
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(italic=True, size=10)
    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")

    # Headers
    row = ws_attacks.max_row + 1
    headers = [
        "Player Name", "Player Tag", "TH Level", "War Date", "War Type",
        "Map Position", "Enemy Name", "Enemy Tag", "Enemy TH", "Enemy Map Position",
        "Stars", "Destruction %", "Attack Order", "Fresh Attack", "War Tag", "War State"
    ]

    for i, header in enumerate(headers, 1):
        cell = ws_attacks.cell(row=row, column=i)
        cell.value = header

    start_row = row

    # Add data rows
    for hit in war_hits:
        row = ws_attacks.max_row + 1
        ws_attacks.append([
            str(hit.get("attacker_name", "")),
            str(hit.get("attacker_tag", "")),
            str(hit.get("attacker_townhall", "")),
            str(hit.get("war_date", "")),
            str(hit.get("war_type", "")),
            str(hit.get("attacker_map_position", "")),
            str(hit.get("defender_name", "")),
            str(hit.get("defender_tag", "")),
            str(hit.get("defender_townhall", "")),
            str(hit.get("defender_map_position", "")),
            hit.get("stars", 0) or 0,
            hit.get("destruction_percentage", 0) or 0,
            str(hit.get("order", "")),
            "Yes" if hit.get("fresh_attack", False) else "No",
            str(hit.get("war_tag", "")),
            str(hit.get("war_state", ""))
        ])

    end_row = ws_attacks.max_row
    format_table(ws_attacks, start_row, end_row, "AttackStats")

    # Now populate the Defense Details sheet
    await insert_logo_from_cdn(
        ws_defenses,
        image_url="https://assets.clashk.ing/logos/crown-text-white-bg/BqlEp974170917vB1qK0zunfANJCGi0W031dTksEq7KQ9LoXWMFk0u77unHJa.png",
        anchor_cell="H1",
        height=50
    )

    # Title for Defense Details
    row = ws_defenses.max_row + 3
    ws_defenses.merge_cells(f"A{row}:P{row}")
    cell = ws_defenses.cell(row=row, column=1)
    cell.value = f"Defense Details for {tag}"
    cell.alignment = Alignment(horizontal="center")
    cell.fill = clashking_theme["title_fill"]
    cell.font = Font(bold=True, size=14)

    # Add filters under title
    row = ws_defenses.max_row + 1
    ws_defenses.merge_cells(f"A{row}:P{row}")
    cell = ws_defenses.cell(row=row, column=1)
    cell.value = get_filter_summary()
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(italic=True, size=10)
    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")

    # Headers for Defense Details
    row = ws_defenses.max_row + 1
    defense_headers = [
        "Defender Name", "Defender Tag", "Defender TH", "War Date", "War Type",
        "Map Position", "Attacker Name", "Attacker Tag", "Attacker TH", "Attacker Map Position",
        "Stars Given", "Destruction Given %", "Attack Order", "Fresh Attack", "War Tag", "War State"
    ]

    for i, header in enumerate(defense_headers, 1):
        cell = ws_defenses.cell(row=row, column=i)
        cell.value = header

    defense_start_row = row

    # Add defense data rows (same data but from defender perspective)
    for hit in war_hits:
        row = ws_defenses.max_row + 1
        ws_defenses.append([
            str(hit.get("defender_name", "")),
            str(hit.get("defender_tag", "")),
            str(hit.get("defender_townhall", "")),
            str(hit.get("war_date", "")),
            str(hit.get("war_type", "")),
            str(hit.get("defender_map_position", "")),
            str(hit.get("attacker_name", "")),
            str(hit.get("attacker_tag", "")),
            str(hit.get("attacker_townhall", "")),
            str(hit.get("attacker_map_position", "")),
            hit.get("stars", 0) or 0,
            hit.get("destruction_percentage", 0) or 0,
            str(hit.get("order", "")),
            "Yes" if hit.get("fresh_attack", False) else "No",
            str(hit.get("war_tag", "")),
            str(hit.get("war_state", ""))
        ])

    defense_end_row = ws_defenses.max_row
    format_table(ws_defenses, defense_start_row, defense_end_row, "DefenseStats")

    # Now populate the Summary sheet (already created as first sheet)

    await insert_logo_from_cdn(
        ws_summary,
        image_url="https://assets.clashk.ing/logos/crown-text-white-bg/BqlEp974170917vB1qK0zunfANJCGi0W031dTksEq7KQ9LoXWMFk0u77unHJa.png",
        anchor_cell="D1",
        height=50
    )

    # Summary statistics
    total_attacks = len(war_hits)
    total_stars = sum(hit.get("stars", 0) for hit in war_hits)
    total_destruction = sum(hit.get("destruction_percentage", 0) for hit in war_hits)
    three_stars = sum(1 for hit in war_hits if hit.get("stars") == 3)
    two_stars = sum(1 for hit in war_hits if hit.get("stars") == 2)
    one_star = sum(1 for hit in war_hits if hit.get("stars") == 1)
    zero_stars = sum(1 for hit in war_hits if hit.get("stars") == 0)

    # Additional statistics
    total_wars = len(set(hit.get("war_tag", "") for hit in war_hits if hit.get("war_tag")))

    # TH level analysis
    th_attacks = {}
    opponent_th_attacks = {}
    th_stars_breakdown = {}  # TH -> {0: count, 1: count, 2: count, 3: count}

    # TH matchup analysis (Attacker TH vs Defender TH)
    attack_matchups = {}  # (attacker_th, defender_th) -> {0: count, 1: count, 2: count, 3: count, total_destruction: float}
    defense_matchups = {}  # (defender_th, attacker_th) -> {0: count, 1: count, 2: count, 3: count, total_destruction: float}

    for hit in war_hits:
        own_th = hit.get("attacker_townhall", "Unknown")
        enemy_th = hit.get("defender_townhall", "Unknown")
        stars = hit.get("stars", 0)
        destruction = hit.get("destruction_percentage", 0)

        th_attacks[own_th] = th_attacks.get(own_th, 0) + 1
        opponent_th_attacks[enemy_th] = opponent_th_attacks.get(enemy_th, 0) + 1

        # Track stars by TH level
        if own_th not in th_stars_breakdown:
            th_stars_breakdown[own_th] = {0: 0, 1: 0, 2: 0, 3: 0}
        th_stars_breakdown[own_th][stars] = th_stars_breakdown[own_th].get(stars, 0) + 1

        # Track attack matchups (when attacking)
        if own_th != "Unknown" and enemy_th != "Unknown":
            matchup_key = (own_th, enemy_th)
            if matchup_key not in attack_matchups:
                attack_matchups[matchup_key] = {0: 0, 1: 0, 2: 0, 3: 0, 'total_destruction': 0, 'attack_count': 0}
            attack_matchups[matchup_key][stars] = attack_matchups[matchup_key].get(stars, 0) + 1
            attack_matchups[matchup_key]['total_destruction'] += destruction
            attack_matchups[matchup_key]['attack_count'] += 1

            # Track defense matchups (when being attacked) - reverse perspective
            defense_key = (enemy_th, own_th)  # defender TH, attacker TH
            if defense_key not in defense_matchups:
                defense_matchups[defense_key] = {0: 0, 1: 0, 2: 0, 3: 0, 'total_destruction': 0, 'attack_count': 0}
            defense_matchups[defense_key][stars] = defense_matchups[defense_key].get(stars, 0) + 1
            defense_matchups[defense_key]['total_destruction'] += destruction
            defense_matchups[defense_key]['attack_count'] += 1

    most_common_th = max(th_attacks.items(), key=lambda x: x[1]) if th_attacks else ("Unknown", 0)
    most_targeted_th = max(opponent_th_attacks.items(), key=lambda x: x[1]) if opponent_th_attacks else ("Unknown", 0)
    
    # Determine current player TH (highest TH level used in attacks)
    current_player_th = max((int(th) for th in th_attacks.keys() if str(th).isdigit()), default=0)

    # Attack performance by stars
    perfect_rate = round(three_stars / total_attacks * 100, 2) if total_attacks > 0 else 0
    fail_rate = round(zero_stars / total_attacks * 100, 2) if total_attacks > 0 else 0

    # Destruction analysis
    high_destruction = sum(1 for hit in war_hits if hit.get("destruction_percentage", 0) >= 90)
    medium_destruction = sum(1 for hit in war_hits if 50 <= hit.get("destruction_percentage", 0) < 90)
    low_destruction = sum(1 for hit in war_hits if hit.get("destruction_percentage", 0) < 50)

    # War type breakdown
    cwl_attacks = sum(1 for hit in war_hits if hit.get("war_type") == "CWL")
    regular_attacks = sum(1 for hit in war_hits if hit.get("war_type") == "Regular")
    friendly_war_attacks = sum(1 for hit in war_hits if hit.get("war_type") not in ["CWL", "Regular"])

    # Fresh vs cleanup attacks (once we fix the detection)
    fresh_attacks = sum(1 for hit in war_hits if hit.get("fresh_attack", False))
    cleanup_attacks = total_attacks - fresh_attacks

    row = ws_summary.max_row + 3
    ws_summary.merge_cells(f"A{row}:B{row}")
    cell = ws_summary.cell(row=row, column=1)
    cell.value = f"War Statistics Summary for {tag}"
    cell.alignment = Alignment(horizontal="center")
    cell.fill = clashking_theme["title_fill"]
    cell.font = Font(bold=True, size=14)

    # Add filters under title on summary sheet too
    row = ws_summary.max_row + 1
    ws_summary.merge_cells(f"A{row}:B{row}")
    cell = ws_summary.cell(row=row, column=1)
    cell.value = get_filter_summary()
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(italic=True, size=10)
    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")

    # We'll create tables instead of including TH breakdown in summary data

    summary_data = [
        # Top Level Stats (moved to top)
        ["Total Wars", int(total_wars)],
        ["Total Attacks", int(total_attacks)],
        ["Total Stars", int(total_stars)],

        ["", ""],  # Separator

        # War Stats
        ["CWL Attacks",
         f"{cwl_attacks} ({round(cwl_attacks / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],
        ["Regular War Attacks",
         f"{regular_attacks} ({round(regular_attacks / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],
        ["Friendly War Attacks",
         f"{friendly_war_attacks} ({round(friendly_war_attacks / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],

        ["", ""],  # Separator

        # Average Stats
        ["Average Stars", round(total_stars / total_attacks, 2) if total_attacks > 0 else 0],
        ["Average Destruction", f"{round(total_destruction / total_attacks, 2)}%" if total_attacks > 0 else "0%"],

        ["", ""],  # Separator

        # Attack Distribution
        ["3 Star Attacks", f"{three_stars} ({perfect_rate}%)"],
        ["2 Star Attacks",
         f"{two_stars} ({round(two_stars / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],
        ["1 Star Attacks",
         f"{one_star} ({round(one_star / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],
        ["0 Star Attacks", f"{zero_stars} ({fail_rate}%)"],

        ["", ""],  # Separator

        # TH Analysis
        ["Most Used TH", f"TH{most_common_th[0]} ({most_common_th[1]} attacks)"],
        ["Most Targeted TH", f"TH{most_targeted_th[0]} ({most_targeted_th[1]} attacks)"],

        ["", ""],  # Separator

        # Destruction Analysis
        ["High Destruction (≥90%)",
         f"{high_destruction} ({round(high_destruction / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],
        ["Medium Destruction (50-89%)",
         f"{medium_destruction} ({round(medium_destruction / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],
        ["Low Destruction (<50%)",
         f"{low_destruction} ({round(low_destruction / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],

        ["", ""],  # Separator

        # Attack Type (when fresh detection is fixed)
        ["Fresh Attacks",
         f"{fresh_attacks} ({round(fresh_attacks / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"],
        ["Cleanup Attacks",
         f"{cleanup_attacks} ({round(cleanup_attacks / total_attacks * 100, 2)}%)" if total_attacks > 0 else "0 (0%)"]

    ]

    ws_summary.append(["Statistic", "Value"])
    start_summary_row = ws_summary.max_row

    for stat_name, stat_value in summary_data:
        ws_summary.append([str(stat_name), str(stat_value)])

    end_summary_row = ws_summary.max_row
    format_table(ws_summary, start_summary_row, end_summary_row, "Summary")
    ws_summary.append([])  # Empty row after Summary table

    # Add TH Attack Performance Table
    if th_stars_breakdown:
        # Table title
        row = ws_summary.max_row + 3
        ws_summary.merge_cells(f"A{row}:F{row}")
        cell = ws_summary.cell(row=row, column=1)
        cell.value = "Attack Performance by Town Hall Level"
        cell.alignment = Alignment(horizontal="center")
        cell.fill = clashking_theme["header_fill_black"]
        cell.font = clashking_theme["header_font_white"]

        # Table headers
        ws_summary.append(["TH Level", "Total Attacks", "3 Stars", "2 Stars", "1 Star", "0 Stars"])
        th_table_start = ws_summary.max_row

        # Track rows that should be highlighted (current player TH)
        current_th_rows = []

        # Sort TH levels numerically
        for th in sorted(th_stars_breakdown.keys(), key=lambda x: int(x) if str(x).isdigit() else 999):
            if th == "Unknown":
                continue
            stars_data = th_stars_breakdown[th]
            total_th_attacks = sum(stars_data.values())
            if total_th_attacks > 0:
                ws_summary.append([
                    f"TH{th}",
                    total_th_attacks,
                    f"{stars_data[3]} ({round(stars_data[3] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[2]} ({round(stars_data[2] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[1]} ({round(stars_data[1] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[0]} ({round(stars_data[0] / total_th_attacks * 100, 1)}%)"
                ])
                
                # Mark current player TH row for highlighting
                if int(th) == current_player_th:
                    current_th_rows.append(ws_summary.max_row)

        th_table_end = ws_summary.max_row
        format_table(ws_summary, th_table_start, th_table_end, "THAttacks", highlight_rows=current_th_rows)
        ws_summary.append([])  # Empty row after TH Attacks table

        # Attack Performance Matrix
        row = ws_summary.max_row + 3
        ws_summary.merge_cells(f"A{row}:I{row}")
        cell = ws_summary.cell(row=row, column=1)
        cell.value = "Attack Performance Matrix"
        cell.alignment = Alignment(horizontal="center")
        cell.fill = clashking_theme["header_fill_black"]
        cell.font = clashking_theme["header_font_white"]

        # Enhanced headers with efficiency indicators and destruction average
        ws_summary.append(["Matchup", "Attacks", "3⭐", "2⭐", "1⭐", "0⭐", "Avg Stars", "Avg Destruction", "Attacks Rate"])
        attack_table_start = ws_summary.max_row

        # Track rows that should be highlighted (current player TH attacks)
        current_th_attack_rows = []

        # Sort by your TH first, then enemy TH
        for (your_th, enemy_th) in sorted(attack_matchups.keys(), key=lambda x: (
                int(x[0]) if str(x[0]).isdigit() else 999, int(x[1]) if str(x[1]).isdigit() else 999)):
            stars_data = attack_matchups[(your_th, enemy_th)]
            # Calculate total attacks excluding metadata keys
            total_matchup_attacks = stars_data.get('attack_count', sum(v for k, v in stars_data.items() if k in [0, 1, 2, 3]))
            if total_matchup_attacks > 0:
                # Calculate average stars
                avg_stars = (stars_data[3] * 3 + stars_data[2] * 2 + stars_data[1] * 1) / total_matchup_attacks

                # Calculate average destruction
                avg_destruction = stars_data.get('total_destruction', 0) / total_matchup_attacks if total_matchup_attacks > 0 else 0

                # Calculate efficiency based on 3 stars being perfect (100%)
                efficiency = f"{(avg_stars / 3.0 * 100):.0f}%"

                ws_summary.append([
                    f"TH{your_th} vs TH{enemy_th}",
                    total_matchup_attacks,
                    f"{stars_data[3]} ({round(stars_data[3] / total_matchup_attacks * 100, 1)}%)",
                    f"{stars_data[2]} ({round(stars_data[2] / total_matchup_attacks * 100, 1)}%)",
                    f"{stars_data[1]} ({round(stars_data[1] / total_matchup_attacks * 100, 1)}%)",
                    f"{stars_data[0]} ({round(stars_data[0] / total_matchup_attacks * 100, 1)}%)",
                    f"{avg_stars:.2f}",
                    f"{avg_destruction:.1f}%",
                    efficiency
                ])
                
                # Mark current player TH rows for highlighting
                if int(your_th) == current_player_th:
                    current_th_attack_rows.append(ws_summary.max_row)

        attack_table_end = ws_summary.max_row
        format_table(ws_summary, attack_table_start, attack_table_end, "AttackMatrix", highlight_rows=current_th_attack_rows)
        ws_summary.append([])  # Empty row after Attack Matrix table

    # Add Defense Performance Matrix (detailed matchups like attacks)
    if defense_matchups:
        # Table title
        row = ws_summary.max_row + 3
        ws_summary.merge_cells(f"A{row}:I{row}")
        cell = ws_summary.cell(row=row, column=1)
        cell.value = "Defense Performance Matrix"
        cell.alignment = Alignment(horizontal="center")
        cell.fill = clashking_theme["header_fill_black"]
        cell.font = clashking_theme["header_font_white"]

        # Detailed defense headers matching attack format
        ws_summary.append(
            ["Matchup", "Defenses", "3⭐", "2⭐", "1⭐", "0⭐", "Avg Stars", "Avg Destruction", "Defense Rate"])
        defense_table_start = ws_summary.max_row

        # Track rows that should be highlighted (current player TH defenses)
        current_th_defense_rows = []

        # Sort by enemy TH first, then your TH
        for (enemy_th, your_th) in sorted(defense_matchups.keys(), key=lambda x: (
                int(x[0]) if str(x[0]).isdigit() else 999, int(x[1]) if str(x[1]).isdigit() else 999)):
            stars_data = defense_matchups[(enemy_th, your_th)]
            # Calculate total attacks excluding metadata keys
            total_defense_attacks = stars_data.get('attack_count', sum(v for k, v in stars_data.items() if k in [0, 1, 2, 3]))
            if total_defense_attacks > 0:
                # Calculate average stars given up
                avg_stars_given = (stars_data[3] * 3 + stars_data[2] * 2 + stars_data[1] * 1) / total_defense_attacks

                # Calculate average destruction given up
                avg_destruction_given = stars_data.get('total_destruction', 0) / total_defense_attacks if total_defense_attacks > 0 else 0

                # Calculate defense rate (0 stars + partial holds)
                held_completely = stars_data[0]
                partial_holds = stars_data[1] + stars_data[2]
                defense_rate = f"{((held_completely + partial_holds) / total_defense_attacks * 100):.1f}%"

                ws_summary.append([
                    f"TH{enemy_th} vs TH{your_th}",
                    total_defense_attacks,
                    f"{stars_data[3]} ({round(stars_data[3] / total_defense_attacks * 100, 1)}%)",
                    f"{stars_data[2]} ({round(stars_data[2] / total_defense_attacks * 100, 1)}%)",
                    f"{stars_data[1]} ({round(stars_data[1] / total_defense_attacks * 100, 1)}%)",
                    f"{stars_data[0]} ({round(stars_data[0] / total_defense_attacks * 100, 1)}%)",
                    f"{avg_stars_given:.2f}",
                    f"{avg_destruction_given:.1f}%",
                    defense_rate
                ])
                
                # Mark current player TH rows for highlighting (defending TH)
                if int(your_th) == current_player_th:
                    current_th_defense_rows.append(ws_summary.max_row)

        defense_table_end = ws_summary.max_row
        format_table(ws_summary, defense_table_start, defense_table_end, "DefenseMatrix", highlight_rows=current_th_defense_rows)
        ws_summary.append([])  # Empty row after Defense Matrix table

    # Save to temporary file
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.seek(0)

    # Generate filename
    player_name = str(war_hits[0].get("attacker_name", "player")) if war_hits else "player"
    # Clean player name for filename
    clean_player_name = "".join(c for c in player_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    clean_tag = tag.replace('#', '')
    filename = f"war_stats_{clean_player_name}_{clean_tag}.xlsx"

    return FileResponse(
        path=tmp.name,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.post("/war/clans/warhits")
async def clan_warhits_stats(filter: ClanWarHitsFilter):
    client = coc.Client(raw_attribute=True)
    START = pend.from_timestamp(filter.timestamp_start, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')
    END = pend.from_timestamp(filter.timestamp_end, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')
    clan_tags = [fix_tag(tag) for tag in filter.clan_tags]

    async def fetch_clan(clan_tag: str):
        pipeline = [
            {"$match": {
                "data.clan.tag": clan_tag,
                "data.preparationStartTime": {"$gte": START, "$lte": END}
            }},
            {"$unset": ["_id"]},
            {"$project": {"data": "$data"}},
            {"$sort": {"data.preparationStartTime": -1}},
            {"$limit": filter.limit or 100},
        ]

        wars_docs = await mongo.clan_wars.aggregate(pipeline, allowDiskUse=True).to_list(length=None)

        results = await collect_player_hits_from_wars(
            wars_docs,
            tags_to_include=None,
            clan_tags=[clan_tag],
            filter=filter,
            client=client,
        )

        return {
            "clan_tag": clan_tag,
            "players": results["items"],
            "wars": results["wars"]
        }

    clan_tasks = [fetch_clan(tag) for tag in clan_tags]
    items = await asyncio.gather(*clan_tasks)

    return {"items": items}
